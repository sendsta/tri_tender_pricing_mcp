from pathlib import Path
from typing import Union

from openpyxl import load_workbook


def read_xlsx_text(path: Union[str, Path]) -> str:
    """
    Read text‑like content from all cells in an XLSX workbook.

    This is intentionally simple—its goal is just to surface enough context
    to the LLM, not to perfectly parse complex BOQs.
    """
    p = Path(path)
    if not p.exists():
        return ""

    try:
        wb = load_workbook(str(p), read_only=True, data_only=True)
    except Exception:
        return ""

    cells = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                vals = [str(v) for v in row if v not in (None, "")]
                if vals:
                    cells.append(" \t ".join(vals))
    finally:
        wb.close()

    return "\n".join(cells)
